
"""
Using an excel file, this program creates a dictionary of celebrities with June birthdays.
The user can add to or delete birthdays from the dictionary. Using the current year, the
program prints the age and birth date of each person in the final dictionary. Finally, a
a full calendar of names and dates is shown using tkinter.
"""
from datetime import datetime
import datetime
import openpyxl
import tkinter

CANVAS_SIZE = 575
N_ROWS = 5
N_COLS = 7
SIZE_ROW = CANVAS_SIZE / N_ROWS - 1
SIZE_COL = CANVAS_SIZE / N_COLS - 1

FILENAME = "June_Birthdays.xlsx"
june_birthdays = {}
days = {0: 'Sun', 1: 'Mon', 2: 'Tues', 3: 'Wed', 4: 'Thurs', 5: 'Fri', 6: "Sat"}

def main():
    make_dictionary_from_excel_file()
    edit_dictionary()
    print_ages()
    show_calendar()

def show_calendar():
    canvas = make_canvas(CANVAS_SIZE, CANVAS_SIZE + 100, 'Calendar')
    for row in range(N_ROWS):
        for col in range(N_COLS):
            draw_square(canvas, row, col)
    labels(canvas)
    canvas.mainloop()

def draw_square(canvas, row, col):
    x = (col * SIZE_COL)
    y = (row * SIZE_ROW)
    canvas.create_rectangle(x+3, y+3, x + SIZE_COL, y + SIZE_ROW, outline='black')

def labels(canvas):
    # month/year label
    canvas.create_text(200, 620, anchor='w', font='Courier 52 bold', text='June 2020', fill='coral')
    # days of the week
    for i in range(7):
        canvas.create_text(15 + i * SIZE_COL, 15, anchor='w', font='Courier 10 bold', text=days[i], fill='coral')

    # first week
    for i in range(1, 8):
        canvas.create_text(73 + i * SIZE_COL, 105, anchor='w', font='Courier 9', text=str(i))
        for key in june_birthdays.keys():
            if june_birthdays[key].day == i:
                canvas.create_text(8 + i * SIZE_COL, 90, anchor='w', font='Courier 7', text=key, fill='blue')
    # middle 3 weeks
    for row in range(2, 5):
        for col in range(N_COLS):
            # day of month to be shown
            canvas.create_text(67 + col * SIZE_COL, row * 111, anchor='w', font='Courier 9', text=str(col+7*(row-1)))
            for key in june_birthdays.keys():
                if june_birthdays[key].day == col+7*(row-1):
                    canvas.create_text(8 + col * SIZE_COL, row * 101, anchor='w', font='Courier 7', text=key,
                                       fill='blue')
    # last week
    for i in range(3):
        canvas.create_text(67 + i * SIZE_COL, 560, anchor='w', font='Courier 9', text=str(i + 28))
        for key in june_birthdays.keys():
            if june_birthdays[key].day == i + 28:
                canvas.create_text(8 + i * SIZE_COL, 545, anchor='w', font='Courier 7', text=key, fill='blue')

    return canvas

def make_dictionary_from_excel_file():
    # reading the excel file
    wb = openpyxl.load_workbook(FILENAME)
    sheet = wb['Sheet1']

    # transferring cell values to dictionary
    for i in range(2, 18):
        name = str(sheet.cell(row=i, column=1).value)
        birth_date = sheet.cell(row=i, column=2).value
        date_without_time = datetime.datetime.date(birth_date)  # gets rid of time (00:00:00)
        june_birthdays[name] = date_without_time

def edit_dictionary():
    # Asking user if they'd like to add or delete from june birthday dictionary
    answer = input("Enter 'A' to add a birthday, 'D' to delete a birthday, or any other key to see ages: ")
    while answer == 'A' or answer == 'D':
        if answer == 'A':
            new_name = input("Enter name: ")
            date_entry = (input("Enter birthday YYYY-MM-DD: "))
            year, month, day = map(int, date_entry.split('-')) # stack overflow
            new_date = datetime.date(year, month, day)
            june_birthdays[new_name] = new_date
            print(str(new_name) + "'s birthday has been added!")
            print(" ")

        if answer == 'D':
            delete_name = input("To delete this person from the birthday calendar, enter their first and last name: ")
            if delete_name not in june_birthdays:
                print("Sorry, this name isn't on the calendar anyway.")
            else:
                del june_birthdays[delete_name]
                print(str(delete_name) + " is not on the birthday calendar anymore!")
                print(" ")

        answer = input("Enter 'A' to add a birthday, 'D' to delete a birthday, or any other key to see ages: ")

def print_ages():
    # print birthdays and ages
    print("June Birthdays")
    for key in june_birthdays.keys():
        current_age = datetime.datetime.now().year - june_birthdays[key].year
        print(str(key) + " will be " + str(current_age) + " years old this year.")
        print(str(june_birthdays[key]))
        print(" ")

def make_canvas(width, height, title=None):
    """
    DO NOT MODIFY
    Creates and returns a drawing canvas
    of the given int size with a blue border,
    ready for drawing.
    """
    objects = {}
    top = tkinter.Tk()
    top.minsize(width=width, height=height)
    if title:
        top.title(title)
    canvas = tkinter.Canvas(top, width=width + 1, height=height + 1)
    canvas.pack()
    canvas.xview_scroll(8, 'units')  # add this so (0, 0) works correctly
    canvas.yview_scroll(8, 'units')  # otherwise it's clipped off

    return canvas

if __name__ == '__main__':
    main()